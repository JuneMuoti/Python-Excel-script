import openpyxl as xl
from openpyxl.chart import BarChart,Reference


def correct_price(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    rows = sheet.max_row
    for row in range(2, rows + 1):
        cell = sheet.cell(row, 3)

        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_col=4, min_row=2, max_col=4, max_row=sheet.max_row)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)
